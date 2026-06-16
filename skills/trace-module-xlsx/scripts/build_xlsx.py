#!/usr/bin/env python3
"""Build a module-breakdown comparison xlsx from a JSON spec.

  python build_xlsx.py spec.json              # write the xlsx
  python build_xlsx.py spec.json --coverage   # only print module + unclassified
                                              # coverage, write nothing (use this
                                              # FIRST on a new model/framework)

Spec schema (see examples/spec_example.json):
{
  "title":  "<workbook title>",
  "output": "/abs/path/out.xlsx",
  "datasets": [
     {"label":"M3 sglang MI355 4k prefill",     # unique; used to reference below
      "model":"minimax-m3",                      # picks the rule set (free-form ok)
      "framework":"sglang", "gpu":"MI355",       # metadata only (shown in header)
      "trace":"/abs/glob-to-ONE-rank*TP-0*.json.gz",
      "region":"prefill",                        # all | prefill | decode
      "avg_ranks": false},                       # true -> average over all TP-* files
     ...
  ],
  "sheets": [
     {"title":"prefill 4k vs 100k",
      "compare":["M3 sglang MI355 4k prefill","M3 sglang MI355 100k prefill"],
      "note":"TOTAL prompt (GPU-busy ms)"},
     ...
  ],
  "detail": true        # also emit one per-kernel detail sheet per dataset
}

Notes baked into output:
- numbers are GPU-busy ms (sum of kernel durations) of the chosen region.
- DECODE region of a CUDA-graphed run = ~ONE graph step in the trace; that matches
  measured wall-clock ms/step. Do NOT divide a decode region by the step count.
- ratio columns are vs the FIRST dataset in each sheet.
"""
import argparse
import glob
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402

from trace_modules import analyze, merged_order  # noqa: E402

BOLD = Font(bold=True)
HDR = Font(bold=True, color="FFFFFF")
HDRFILL = PatternFill("solid", fgColor="305496")
SUBFILL = PatternFill("solid", fgColor="DDEBF7")
TITLEFILL = PatternFill("solid", fgColor="1F4E78")
TITLEFONT = Font(bold=True, color="FFFFFF", size=12)


def _short(name, n=31):
    return name[:n]


def _resolve_source(ds):
    """Pick the trace source for a dataset.

    PREFERENCE: a profile_by_stage file (already one phase, region='all', exact
    boundary) over slicing a combined trace.
      - stage_trace  : glob to the by_stage file for this phase (preferred).
                       prefill -> the '*-EXTEND*' file; decode -> the no-suffix file.
      - combined_trace + phase : fallback; sliced by phase via decode markers.
      - trace + region         : legacy explicit form.
    Returns (trace_glob, region, source_str).
    """
    st = ds.get("stage_trace")
    if st and glob.glob(st):
        return st, "all", "by_stage"
    if ds.get("combined_trace"):
        phase = ds.get("phase")
        if phase not in ("prefill", "decode"):
            raise ValueError(f"{ds['label']}: combined_trace needs phase=prefill|decode")
        tag = "combined-slice(fallback)" if st else "combined-slice"
        return ds["combined_trace"], phase, tag
    if ds.get("trace"):
        return ds["trace"], ds.get("region", "all"), "explicit"
    raise ValueError(f"{ds['label']}: provide stage_trace, combined_trace+phase, or trace")


def build(spec, coverage_only=False):
    cache = {}
    disp = {}     # unique label -> header display name (defaults to the label)
    source = {}   # label -> which source was used
    for ds in spec["datasets"]:
        trace, region, src = _resolve_source(ds)
        cache[ds["label"]] = analyze(
            trace, ds.get("model", "generic"),
            region=region, avg_ranks=ds.get("avg_ranks", False),
        )
        disp[ds["label"]] = ds.get("display", ds["label"])
        source[ds["label"]] = src

    print("=== coverage report ===")
    for label, a in cache.items():
        u = a["modules"].get("other/unclassified", 0.0)
        pct = 100 * u / a["total_ms"] if a["total_ms"] else 0
        flag = "  <-- HIGH, extend ruleset" if pct > 3 else ""
        msg = f"  [{a['ruleset']}/{source[label]}] {label}: total {a['total_ms']:.1f} ms, unclassified {u:.1f} ms ({pct:.1f}%){flag}"
        if a["note"]:
            msg += f"  [{a['note']}]"
        print(msg)
    if coverage_only:
        # also dump top unclassified kernels to help extend rules
        for label, a in cache.items():
            unc = a["kernels"].get("other/unclassified", {})
            if not unc:
                continue
            print(f"\n  top unclassified in '{label}':")
            for name, (c, ms) in sorted(unc.items(), key=lambda x: -x[1][1])[:12]:
                print(f"    {ms:8.2f} ms  {c:5d}x  {name[:70]}")
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet in spec["sheets"]:
        labels = sheet["compare"]
        ws = wb.create_sheet(_short(sheet["title"]))
        order = merged_order([cache[l]["order"] for l in labels])
        # title
        meta = " | ".join(
            f"{disp[l]}=[{_ds(spec,l).get('framework','?')}/{_ds(spec,l).get('gpu','?')}]"
            for l in labels
        )
        ws.append([f"{spec.get('title','')}  —  {sheet['title']}   ({meta})"])
        ncol = 1 + 2 * len(labels) + (len(labels) - 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ncol, 2))
        ws["A1"].fill = TITLEFILL
        ws["A1"].font = TITLEFONT
        if sheet.get("note"):
            ws.append([sheet["note"]])
        # header
        hdr = ["module"]
        for l in labels:
            hdr += [f"{disp[l]} ms", f"{disp[l]} %"]
        base = labels[0]
        for l in labels[1:]:
            hdr.append(f"x({disp[l]}/{disp[base]})")
        ws.append(hdr)
        for c in ws[ws.max_row]:
            c.font = HDR
            c.fill = HDRFILL
        # rows
        tots = {l: cache[l]["total_ms"] for l in labels}
        for mod in order:
            vals = {l: cache[l]["modules"].get(mod, 0.0) for l in labels}
            if all(v == 0 for v in vals.values()):
                continue
            row = [mod]
            for l in labels:
                t = tots[l] or 1
                row += [round(vals[l], 2), round(100 * vals[l] / t, 1)]
            for l in labels[1:]:
                row.append(round(vals[l] / vals[base], 2) if vals[base] > 0 else "n/a")
            ws.append(row)
        # total
        trow = ["TOTAL"]
        for l in labels:
            trow += [round(tots[l], 2), None]
        for l in labels[1:]:
            trow.append(round(tots[l] / tots[base], 2) if tots[base] > 0 else "n/a")
        ws.append(trow)
        for c in ws[ws.max_row]:
            c.font = BOLD
            c.fill = SUBFILL
        ws.column_dimensions["A"].width = 42
        for i in range(1, ncol + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = 13

    if spec.get("detail", True):
        seen = set()
        for sheet in spec["sheets"]:
            nm = _short("det " + sheet["title"])
            while nm in seen:
                nm = _short(nm[:29] + "_")
            seen.add(nm)
            _detail_compare(wb.create_sheet(nm), sheet["compare"], cache, disp)

    out = spec["output"]
    wb.save(out)
    print(f"\nwrote {out}  ({len(spec['sheets'])} comparison sheet(s))")
    return out


def _ds(spec, label):
    for d in spec["datasets"]:
        if d["label"] == label:
            return d
    return {}


def _detail_rows(a):
    """Per-kernel rows for one dataset: [module, kernel, ms, %, calls] grouped by
    module (sorted by ms desc within module) with a subtotal row per module."""
    tot = a["total_ms"] or 1
    rows = []  # (cells, is_subtotal)
    for mod in a["order"]:
        if mod not in a["kernels"]:
            continue
        sub = 0.0
        for name, (calls, ms) in sorted(a["kernels"][mod].items(), key=lambda x: -x[1][1]):
            rows.append(([mod, name[:80], round(ms, 3), round(100 * ms / tot, 2), calls], False))
            sub += ms
        rows.append(([f"-- {mod} subtotal --", None, round(sub, 3), round(100 * sub / tot, 2), None], True))
    return rows


def _detail_compare(ws, labels, cache, disp=None):
    """Per-kernel detail for several datasets laid out SIDE BY SIDE (one 5-col block
    each, blank column between), e.g. 4k on the left and 100k on the right."""
    disp = disp or {l: l for l in labels}
    BLK = 5
    ncol = BLK * len(labels) + (len(labels) - 1)
    ws.append([f"per-kernel by module (GPU-busy ms, rank0)   |   " +
               "   ".join(f"BLOCK {i+1}: {disp[l]}" for i, l in enumerate(labels))])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ncol, 2))
    ws["A1"].fill = TITLEFILL
    ws["A1"].font = TITLEFONT
    # header
    hdr = []
    for i, l in enumerate(labels):
        hdr += [f"module ({disp[l]})", "kernel", "ms", "%", "calls"]
        if i < len(labels) - 1:
            hdr += [None]
    ws.append(hdr)
    for c in ws[ws.max_row]:
        if c.value is not None:
            c.font = BOLD
    # body: independent row lists per dataset, padded to the longest
    blocks = [_detail_rows(cache[l]) for l in labels]
    nrows = max(len(b) for b in blocks)
    for r in range(nrows):
        line, subflags = [], []
        for i, b in enumerate(blocks):
            cells, is_sub = b[r] if r < len(b) else ([None] * BLK, False)
            line += cells
            subflags.append(is_sub)
            if i < len(blocks) - 1:
                line += [None]
        ws.append(line)
        for i, is_sub in enumerate(subflags):
            if is_sub:
                base = i * (BLK + 1)
                for c in ws[ws.max_row][base:base + BLK]:
                    c.font = BOLD
                    c.fill = SUBFILL
    # column widths
    for i in range(len(labels)):
        base = i * (BLK + 1) + 1
        for off, w in enumerate((34, 56, 9, 6, 6)):
            ws.column_dimensions[openpyxl.utils.get_column_letter(base + off)].width = w


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("spec")
    ap.add_argument("--coverage", action="store_true", help="print coverage only, write nothing")
    args = ap.parse_args()
    with open(args.spec) as f:
        spec = json.load(f)
    build(spec, coverage_only=args.coverage)


if __name__ == "__main__":
    main()
